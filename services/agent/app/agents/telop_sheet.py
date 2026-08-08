"""Fill a station's or subtitle house's own telop order sheet.

Every programme uses a different sheet, so CurrentCut does not impose a layout.
The director uploads the sheet their programme already uses; Gemini reads its
structure once; CurrentCut then writes the drafted telops into a *copy of that
file*, leaving the formatting, column widths, print area and logo untouched.

Real sheets are not one row per entry. The industry template published by
字幕テロップセンター, for example, gives each entry **two rows** (start time on
the first, end time on the second), puts the caption text in a merged block
spanning both rows, and continues onto a new worksheet every ten entries. The
mapping below therefore describes a repeating *block*, not a set of columns.
"""
from __future__ import annotations

import math
import re
import shutil
import unicodedata
from pathlib import Path

from pydantic import BaseModel, Field

from .. import lang
from ..clients.gemini_client import gemini
from ..models.schemas import TelopEntry

FIELDS = {
    "order": "通し番号 / No.",
    "in_tc": "開始タイム（IN点）",
    "out_tc": "終了タイム（OUT点）",
    "telop_type": "テロップ種別",
    "text": "表示文字（本文）",
    "text_target": "訳文・ターゲット言語の表示文字",
    "source_note": "出典表記",
    "caution": "備考・確認事項",
}


class FieldCell(BaseModel):
    column: str = Field(description="Column letter, e.g. 'H'")
    row_offset: int = Field(default=0, description="Rows below the entry's first row")


class SheetMapping(BaseModel):
    first_data_row: int = Field(description="1-based row of the first entry's first row")
    row_stride: int = Field(default=1, description="How many rows one entry occupies")
    entries_per_sheet: int = Field(default=0, description="0 when the sheet is not paginated")
    sheet_names: list[str] = Field(default_factory=list,
                                   description="Worksheets to fill, in order")
    fields: dict[str, FieldCell] = Field(default_factory=dict)
    max_chars_per_line: int = Field(default=0, description="Limit stated on the sheet, else 0")
    notes: str = ""


_MAPPING_PROMPT = """You are reading a Japanese television telop / subtitle order
sheet (編集指示書・テロップ発注表) so that it can be filled in automatically.

Worksheets in this workbook, in order: {sheets}

Non-empty cells of the first worksheet:
{grid}

Merged cell ranges (these matter — an entry's text often occupies one merged
block spanning several rows and columns):
{merges}

Work out the repeating structure of ONE entry:
- first_data_row: the row where entry 1 begins
- row_stride: how many rows one entry occupies. If a start time and an end time
  sit on consecutive rows for the same entry, the stride is 2, not 1.
- entries_per_sheet: how many entries fit before the sheet is full (0 if it just
  continues down). Worksheet names like "1-10", "11-20" mean ten per sheet.
- sheet_names: the worksheets to fill, in order
- fields: for each of our fields, the column letter and how many rows below the
  entry's first row it sits. For a merged block, give its TOP-LEFT cell.
  Omit any field the sheet has no place for. Never guess.
- max_chars_per_line: if the sheet states a character limit per line
  (e.g. 「1行20文字」), give that number, else 0.

Our fields:
{fields}

Guidance: 「開始タイム」is in_tc and 「終了タイム」is out_tc — usually the same
column, one row apart. 「ソース言語」「元言語」「表示文字」「スーパー」is text.
「ターゲット言語」「訳文」is text_target. 「出典」is source_note. 「備考」is
caution. Ignore columns marked 弊社使用欄 or otherwise reserved for the vendor.

Return JSON."""


def _grid_and_merges(xlsx_path: Path, max_rows: int = 30, max_cols: int = 30):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    cells = []
    for row in range(1, min(ws.max_row, max_rows) + 1):
        for col in range(1, min(ws.max_column, max_cols) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip():
                text = re.sub(r"\s+", " ", str(value).strip())[:44]
                cells.append(f"{get_column_letter(col)}{row}: {text}")
    merges = sorted(str(m) for m in ws.merged_cells.ranges)
    return "\n".join(cells), "\n".join(merges[:60]), wb.sheetnames


def infer_mapping(xlsx_path: str | Path) -> SheetMapping:
    grid, merges, sheetnames = _grid_and_merges(Path(xlsx_path))
    if gemini.mock:
        return _mock_mapping(grid, merges, sheetnames)
    fields = "\n".join(f"- {k}: {v}" for k, v in FIELDS.items())
    try:
        mapping = gemini.structured(
            _MAPPING_PROMPT.format(sheets=", ".join(sheetnames), grid=grid,
                                   merges=merges, fields=fields),
            SheetMapping,
        )
    except Exception as exc:
        raise RuntimeError(f"could not read the sheet layout: {exc}") from exc
    if not mapping.fields:
        raise RuntimeError("no fields could be identified in this sheet")
    mapping.sheet_names = [s for s in mapping.sheet_names if s in sheetnames] or sheetnames
    mapping.row_stride = max(1, mapping.row_stride)
    return mapping


def _mock_mapping(grid: str, merges: str, sheetnames: list[str]) -> SheetMapping:
    """Deterministic structure detection, for tests and for running without
    credentials. Finds the header labels, then derives the block from them."""
    cells: dict[str, tuple[int, str]] = {}   # "B4" -> (row, text)
    for entry in grid.split("\n"):
        m = re.match(r"([A-Z]+)(\d+): (.*)", entry)
        if m:
            cells[m.group(1) + m.group(2)] = (int(m.group(2)), m.group(3))

    keywords = {
        "in_tc": ("開始タイム", "in点", "イン点", "開始"),
        "out_tc": ("終了タイム", "out点", "アウト点", "終了"),
        "text": ("ソース言語", "元言語", "表示文字", "テロップ", "スーパー", "文字"),
        "text_target": ("ターゲット言語", "訳文"),
        "order": ("no", "番号"),
        "telop_type": ("種別", "種類"),
        "source_note": ("出典", "ソース元"),
        "caution": ("備考", "注意", "確認"),
    }
    def is_note(text: str) -> bool:
        """Sheets carry instructions to the person filling them in ("※1行20文字
        …"). Those contain the same words as the headers and must not be
        mistaken for one."""
        return text.startswith(("※", "*", "注")) or "目安" in text or "下さい" in text \
            or "ください" in text or len(text) > 24

    header: dict[str, tuple[str, int]] = {}   # field -> (column, header row)
    for ref, (row, text) in cells.items():
        if "弊社" in text or is_note(text):
            continue
        low = text.lower()
        col = re.match(r"([A-Z]+)", ref).group(1)
        for field, words in keywords.items():
            if field not in header and any(w in low for w in words):
                header[field] = (col, row)

    header_rows = [r for _, r in header.values()] or [1]
    first_data_row = max(header_rows) + 1
    # Two header labels stacked in the same column (start time above end time)
    # mean the entry occupies two rows.
    stride = 1
    if "in_tc" in header and "out_tc" in header:
        (c_in, r_in), (c_out, r_out) = header["in_tc"], header["out_tc"]
        if c_in == c_out and r_out != r_in:
            stride = 2
            first_data_row = max(r_in, r_out) + 1

    fields = {}
    for field, (col, _) in header.items():
        offset = 1 if (field == "out_tc" and stride > 1) else 0
        fields[field] = FieldCell(column=col, row_offset=offset)

    per_sheet = 0
    m = re.match(r"^(\d+)-(\d+)$", sheetnames[0] if sheetnames else "")
    if m:
        per_sheet = int(m.group(2)) - int(m.group(1)) + 1

    limit = 0
    lm = re.search(r"1行\s*(\d+)\s*文字", grid)
    if lm:
        limit = int(lm.group(1))

    return SheetMapping(first_data_row=first_data_row, row_stride=stride,
                        entries_per_sheet=per_sheet, sheet_names=sheetnames,
                        fields=fields, max_chars_per_line=limit,
                        notes="structure detected without an LLM")


# Pouring into a programme's own uploaded form, the wording follows the form,
# not the shoot: a Japanese broadcaster's template expects 名前スーパー in the
# type column whatever language the interview was in. One definition, shared
# with the sheet CurrentCut writes itself.
_TYPE_JA = lang.SHEET_TELOP_TYPE[lang.JA]


def _tc(seconds: float, fps: int = 30, frames: bool = True) -> str:
    seconds = max(0.0, seconds)
    total = int(round(seconds * fps))
    f, s = total % fps, (total // fps) % 60
    m, h = (total // (fps * 60)) % 60, total // (fps * 3600)
    return f"{h:02d}:{m:02d}:{s:02d}:{f:02d}" if frames else f"{h:02d}:{m:02d}:{s:02d}"


def _anchor(ws, column: str, row: int) -> str:
    """A value written to a merged cell must go to its top-left anchor."""
    ref = f"{column}{row}"
    for rng in ws.merged_cells.ranges:
        if ref in rng:
            return rng.coord.split(":")[0]
    return ref


def fill_sheet(template_path: str | Path, mapping: SheetMapping,
               entries: list[TelopEntry], out_path: str | Path) -> Path:
    """Write the entries into a copy of the station's own workbook."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    template_path, out_path = Path(template_path), Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, out_path)

    wb = load_workbook(out_path)
    sheets = [wb[n] for n in mapping.sheet_names if n in wb.sheetnames] or [wb.worksheets[0]]
    per_sheet = mapping.entries_per_sheet or len(entries)
    # Some sheets write times as HH：MM：SS with no frame field; follow the sheet.
    wants_frames = ":" in str(sheets[0][_anchor(sheets[0], *_first_tc_ref(mapping))].value or "")

    written = 0
    for index, entry in enumerate(entries):
        sheet_index, slot = divmod(index, per_sheet)
        if sheet_index >= len(sheets):
            break  # ran out of pages; the remainder is reported to the caller
        ws = sheets[sheet_index]
        base_row = mapping.first_data_row + slot * mapping.row_stride

        values = {
            "order": index + 1,
            "in_tc": _tc(entry.in_seconds, frames=wants_frames),
            "out_tc": _tc(entry.out_seconds, frames=wants_frames),
            "telop_type": _TYPE_JA.get(entry.telop_type, entry.telop_type),
            "text": "\n".join(entry.text_lines),
            "text_target": "",
            "source_note": entry.source_note,
            "caution": entry.caution,
        }
        for field, cell in mapping.fields.items():
            value = values.get(field)
            if value in (None, ""):
                continue
            ref = _anchor(ws, cell.column, base_row + cell.row_offset)
            target = ws[ref]
            target.value = value
            if field in ("text", "text_target") and "\n" in str(value):
                target.alignment = Alignment(wrap_text=True, vertical="center")
        written += 1

    wb.save(out_path)
    if written < len(entries):
        raise OverflowError(
            f"the sheet holds {written} of {len(entries)} entries — add pages to the template")
    return out_path


def _first_tc_ref(mapping: SheetMapping) -> tuple[str, int]:
    cell = mapping.fields.get("in_tc")
    if cell is None:
        return ("A", mapping.first_data_row)
    return (cell.column, mapping.first_data_row + cell.row_offset)


# The columns whose text wraps, and so decide how tall a row has to be.
_WRAPPING_COLUMNS = (5, 6, 7, 8)


def _display_width(text: str) -> int:
    """Columns a string occupies on screen, counting CJK glyphs as two."""
    return sum(2 if unicodedata.east_asian_width(ch) in "WF" else 1 for ch in text)


def _wrapped_lines(text: str, column_width: float) -> int:
    """Lines `text` takes in a wrapped cell of this width."""
    if not text:
        return 1
    usable = max(1.0, column_width - 1)  # the cell keeps a little padding
    return sum(max(1, math.ceil(_display_width(para) / usable))
               for para in str(text).split("\n"))


def write_manuscript(entries: list[TelopEntry], out_path: str | Path,
                     title: str = "", air_date: str = "",
                     language: str = lang.JA) -> Path:
    """The director's own caption order sheet — the file emailed to the edit house.

    This is the everyday deliverable. The programme's template is filled in at
    the edit house, so what the director sends is a plain, readable list: what
    the telop says, when it comes up, and — the part CurrentCut adds — whether
    the figure in it has been checked, and against what.

    Written in the language of the shoot. This is the one deliverable a person
    has to act on line by line, and a column headed 裏付け is unreadable to a
    crew that does not work in Japanese, exactly as `Checked against` would be
    to one that does. The vocabulary lives in app/lang.py with everything else
    that changes by language.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    sheet_title = lang.sheet(lang.SHEET_TITLE, language)
    columns = lang.sheet(lang.SHEET_COLUMNS, language)
    types = lang.sheet(lang.SHEET_TELOP_TYPE, language)
    evidence = lang.sheet(lang.SHEET_EVIDENCE, language)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws["A1"] = sheet_title
    ws["A1"].font = Font(size=14, bold=True)
    programme = lang.sheet(lang.SHEET_PROGRAMME, language)
    air = lang.sheet(lang.SHEET_AIR_DATE, language)
    # Each of these needs room to be read: the programme name ran into the air
    # date and printed as "Programme: The corr".
    ws["A2"] = f"{programme}: {title}" if title else f"{programme}:"
    ws.merge_cells("A2:E2")
    ws["F2"] = f"{air}: {air_date}" if air_date else f"{air}:"
    ws["G2"] = lang.sheet(lang.SHEET_WARNING, language)
    ws.merge_cells("G2:H2")
    ws["G2"].font = Font(size=9, color="8A5A12")

    header_row = 4
    thin = Side(style="thin", color="D0CCC2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EFEBE2")

    for index, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=index, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(index)].width = width

    for offset, entry in enumerate(entries):
        row = header_row + 1 + offset
        values = [
            entry.order,
            _tc(entry.in_seconds),
            _tc(entry.out_seconds),
            types.get(entry.telop_type, entry.telop_type),
            "\n".join(entry.text_lines),
            entry.source_note,
            evidence.get(entry.evidence_status.value, entry.evidence_status.value),
            entry.caution,
        ]
        for index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=index, value=value)
            cell.border = border
            cell.alignment = Alignment(
                wrap_text=index in _WRAPPING_COLUMNS, vertical="center",
                horizontal="center" if index in (1, 2, 3, 7) else "left")
        # Draw the eye to the rows that still need a decision.
        if entry.evidence_status.value == "CONFLICTING":
            ws.cell(row=row, column=7).font = Font(bold=True, color="96311F")
        elif entry.evidence_status.value == "UNVERIFIED":
            ws.cell(row=row, column=7).font = Font(color="8A5A12")

        # Height follows the tallest wrapped cell, not the caption alone. The
        # rows carrying a reason to check — the ones this sheet exists for —
        # have the longest notes, so sizing by the caption hid exactly them.
        tallest = max(_wrapped_lines(str(values[column - 1] or ""),
                                     columns[column - 1][1])
                      for column in _WRAPPING_COLUMNS)
        ws.row_dimensions[row].height = max(20, tallest * 14.5 + 5)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.print_title_rows = f"{header_row}:{header_row}"

    # A caption and the evidence behind it have to land on the same printed
    # page. At eight columns the default portrait width breaks the sheet into
    # column bands, and `Checked against` prints on a page of its own — a list
    # of verdicts with nothing to say which caption each one belongs to. Fit
    # the width to one page and let the rows run on for as many as they need.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    wb.save(out_path)
    return out_path


def write_csv(entries: list[TelopEntry], out_path: str | Path,
              language: str = lang.JA) -> Path:
    """Plain text alternative for anyone who would rather not open a workbook."""
    import csv

    columns = lang.sheet(lang.SHEET_COLUMNS, language)
    types = lang.sheet(lang.SHEET_TELOP_TYPE, language)
    evidence = lang.sheet(lang.SHEET_EVIDENCE, language)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([c[0] for c in columns])
        for e in entries:
            writer.writerow([e.order, _tc(e.in_seconds), _tc(e.out_seconds),
                             types.get(e.telop_type, e.telop_type),
                             "\n".join(e.text_lines), e.source_note,
                             evidence.get(e.evidence_status.value, e.evidence_status.value),
                             e.caution])
    return out_path
