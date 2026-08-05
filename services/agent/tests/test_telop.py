"""Telop drafting and filling a station's own order sheet."""
from pathlib import Path


def test_telops_drafted_with_types_and_timecodes(overnight_run):
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    assert entries, "telops must be drafted"
    assert entries == sorted(entries, key=lambda e: e.order)
    for e in entries:
        assert e.text_lines and any(l.strip() for l in e.text_lines)
        assert e.out_seconds >= e.in_seconds
        assert len(e.text_lines) <= 2, "a telop may not exceed two lines"


def test_telops_carry_no_japanese_punctuation(overnight_run):
    """Broadcast telops do not carry 。 or 、."""
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    for e in store.list(project_id, "telops", TelopEntry):
        joined = "".join(e.text_lines)
        assert "。" not in joined and "、" not in joined, f"punctuation in telop: {joined}"


def test_no_figure_reaches_the_sheet_bare(overnight_run):
    """Every figure on screen carries either its source or a reason it has none.

    Not "confirmed implies attributed" — a claim can be checked and still have
    nobody worth naming on air (see test_citation.py). What must never happen is
    a number arriving on the operator's sheet with neither an attribution nor a
    note about what is missing: that is the one state that reads as settled when
    it is not.
    """
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    data = [e for e in store.list(project_id, "telops", TelopEntry)
            if e.telop_type == "data"]
    assert data, "the demo footage contains figures; some must reach the sheet"
    for e in data:
        assert e.source_note or e.caution, (
            f"figure with neither source nor caveat: {''.join(e.text_lines)}")


def test_no_off_record_text_reaches_the_telop_sheet(overnight_run):
    from app.models.schemas import Segment, TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    off = [s for s in store.list(project_id, "segments", Segment) if "オフレコ" in s.transcript]
    assert off
    for e in store.list(project_id, "telops", TelopEntry):
        joined = "".join(e.text_lines) + e.source_note
        assert "銀座" not in joined
        assert "オフレコ" not in joined


def test_fills_a_station_sheet_keeping_its_own_layout(overnight_run, tmp_path):
    """The filled sheet must be the station's workbook, not a lookalike."""
    from openpyxl import Workbook, load_workbook

    from app.agents import telop_sheet
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)

    # A fictional programme's sheet: title row, then headers on row 3.
    template = tmp_path / "station_format.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "テロップ発注"
    ws["A1"] = "◯◯番組 テロップ発注表"
    ws["A3"], ws["B3"], ws["C3"] = "No", "開始タイム", "終了タイム"
    ws["D3"], ws["E3"], ws["F3"], ws["G3"] = "種別", "表示文字", "出典", "備考"
    ws.column_dimensions["E"].width = 42          # station's own formatting
    wb.save(template)

    mapping = telop_sheet.infer_mapping(template)
    assert mapping.fields, "fields must be identified"
    assert "text" in mapping.fields

    out = telop_sheet.fill_sheet(template, mapping, entries, tmp_path / "filled.xlsx")
    assert out.exists()

    filled = load_workbook(out)
    ws2 = filled["テロップ発注"]
    assert ws2["A1"].value == "◯◯番組 テロップ発注表", "the station's own header must survive"
    assert ws2.column_dimensions["E"].width == 42, "the station's formatting must survive"

    cell = mapping.fields["text"]
    first = ws2[f"{cell.column}{mapping.first_data_row + cell.row_offset}"].value
    assert first and str(first).strip(), "entries must be written into the sheet"


def test_reads_a_real_industry_sheet_structure(tmp_path):
    """The published 編集指示書 gives each entry two rows, puts the text in a
    merged block, and paginates every ten entries. A one-row-per-entry
    assumption silently writes into the wrong cells."""
    from openpyxl import Workbook

    from app.agents import telop_sheet

    template = tmp_path / "real_like.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "1-10"
    ws["A2"] = "※表示字幕文字数の目安は、1行20文字、2行までが読みやすい目安となります"
    ws["B4"], ws["B5"] = "開始タイム", "終了タイム"
    ws["H4"] = "ソース言語（元言語）"
    ws["R4"] = "ターゲット言語（テロップ表示言語）"
    for i in range(10):
        r = 6 + i * 2
        ws.merge_cells(f"A{r}:A{r+1}")
        ws[f"A{r}"] = i + 1
        ws.merge_cells(f"H{r}:Q{r+1}")
    wb.create_sheet("11-20")
    wb.save(template)

    mapping = telop_sheet.infer_mapping(template)
    assert mapping.row_stride == 2, "an entry spans two rows on this sheet"
    assert mapping.first_data_row == 6
    assert mapping.entries_per_sheet == 10, "sheet names like 1-10 mean ten per page"
    assert mapping.max_chars_per_line == 20, "the sheet states its own line limit"
    assert mapping.fields["text"].column == "H", "the note row must not be mistaken for a header"
    assert mapping.fields["in_tc"].row_offset == 0
    assert mapping.fields["out_tc"].row_offset == 1


def test_figures_are_never_broken_across_lines():
    """"5万6000店" split as "5万600" / "0店" is a different number on screen."""
    from app.agents.telop import _fit

    for text in ["コンビニ 全国 5万6000店 手軽に飲める時代",
                 "訪日外国人は356万人 過去最高を更新した",
                 "この店は1日およそ100杯 10年で3割減"]:
        lines = _fit(text)
        for line in lines[:-1]:
            assert not line[-1].isdigit(), f"a line ends mid-figure: {lines}"
        for line in lines[1:]:
            assert not line[0].isdigit() or "万" not in "".join(lines[:1]), lines


def test_manuscript_excel_needs_no_template(overnight_run, tmp_path):
    """The everyday deliverable: the director's テロップ原稿, emailed to the edit
    house, which pours it into the programme's own form. No template required."""
    from openpyxl import load_workbook

    from app.agents import telop_sheet
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    out = telop_sheet.write_manuscript(entries, tmp_path / "原稿.xlsx",
                                       title="トレンド特集", air_date="金曜")
    ws = load_workbook(out)["テロップ原稿"]

    headers = [ws.cell(row=4, column=c).value for c in range(1, 9)]
    assert headers[:5] == ["No", "IN点", "OUT点", "種別", "表示文字"]
    assert "裏付け" in headers, "the edit house must see which figures were checked"
    assert "トレンド特集" in str(ws["A2"].value)

    first_text = ws.cell(row=5, column=5).value
    assert first_text and str(first_text).strip()
    assert ws.freeze_panes is not None, "the header must stay visible when scrolling"

    # Every drafted telop reaches the manuscript.
    rows = sum(1 for r in range(5, 5 + len(entries)) if ws.cell(row=r, column=1).value)
    assert rows == len(entries)


def test_manuscript_marks_figures_that_need_a_decision(overnight_run, tmp_path):
    from openpyxl import load_workbook

    from app.agents import telop_sheet
    from app.models.schemas import EvidenceStatus, TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    out = telop_sheet.write_manuscript(entries, tmp_path / "原稿2.xlsx")
    ws = load_workbook(out)["テロップ原稿"]

    for offset, entry in enumerate(entries):
        label = ws.cell(row=5 + offset, column=7).value
        assert label, "every row states its evidence position"
        if entry.evidence_status == EvidenceStatus.CONFLICTING:
            assert "⚠" in str(label)
        if entry.evidence_status == EvidenceStatus.UNVERIFIED:
            assert "裏付けなし" in str(label)


def test_csv_fallback_when_no_template(overnight_run, tmp_path):
    from app.agents import telop_sheet
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    path = telop_sheet.write_csv(entries, tmp_path / "telops.csv")
    body = Path(path).read_text(encoding="utf-8-sig")
    assert "表示文字" in body and "出典表記" in body
    assert body.count("\n") >= len(entries)


# ---- the sheet is written in the language of the shoot ----------------------
#
# The order sheet is the one deliverable a person acts on line by line. A crew
# that does not read Japanese cannot act on a column headed 裏付け, and the
# column that says whether a figure was checked is the whole point of it.

def test_the_sheet_follows_the_shoot_into_english(overnight_run, tmp_path):
    from openpyxl import load_workbook

    from app import lang
    from app.agents import telop_sheet
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    out = telop_sheet.write_manuscript(entries, tmp_path / "sheet_en.xlsx",
                                       language=lang.EN)
    ws = load_workbook(out)["Caption Order Sheet"]

    headers = [c.value for c in ws[4]]
    assert headers[6] == "Checked against", "the column the product exists for"
    assert headers == ["No", "In", "Out", "Type", "On screen", "Source line",
                       "Checked against", "Notes / to confirm"]
    assert not any("裏付け" in str(h) for h in headers)


def test_evidence_positions_read_in_english_too(overnight_run, tmp_path):
    """A status nobody can read is a status nobody acts on."""
    from openpyxl import load_workbook

    from app import lang
    from app.agents import telop_sheet
    from app.models.schemas import EvidenceStatus, TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    out = telop_sheet.write_manuscript(entries, tmp_path / "sheet_en2.xlsx",
                                       language=lang.EN)
    ws = load_workbook(out)["Caption Order Sheet"]

    for offset, entry in enumerate(entries):
        label = str(ws.cell(row=5 + offset, column=7).value or "")
        assert label and label.isascii() or "⚠" in label, f"row {offset} reads {label!r}"
        if entry.evidence_status == EvidenceStatus.UNVERIFIED:
            assert "not backed" in label


def test_japanese_remains_the_original_wording(overnight_run, tmp_path):
    """The Japanese is the trade's own vocabulary, not a translation back."""
    from openpyxl import load_workbook

    from app import lang
    from app.agents import telop_sheet
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    out = telop_sheet.write_manuscript(entries, tmp_path / "sheet_ja.xlsx",
                                       language=lang.JA)
    ws = load_workbook(out)["テロップ原稿"]
    assert [c.value for c in ws[4]] == ["No", "IN点", "OUT点", "種別", "表示文字",
                                        "出典表記", "裏付け", "備考・確認事項"]


def test_the_csv_follows_the_same_language(overnight_run, tmp_path):
    from app import lang
    from app.agents import telop_sheet
    from app.models.schemas import TelopEntry
    from app.storage import store

    project_id, _ = overnight_run
    entries = store.list(project_id, "telops", TelopEntry)
    body = Path(telop_sheet.write_csv(entries, tmp_path / "en.csv",
                                      language=lang.EN)).read_text(encoding="utf-8-sig")
    header = body.splitlines()[0]
    assert "Checked against" in header and "On screen" in header
    assert "裏付け" not in header
    # Only the sheet's own vocabulary is chosen here. Each row's note was
    # written when the telop was drafted, in the language of that shoot, and it
    # is that shoot's content — this fixture is a Japanese one, so Japanese
    # notes under English headings is the correct outcome rather than a leak.
    # The parameter says which shoot this is, not which language to translate to.


def test_a_programme_template_keeps_its_own_vocabulary():
    """Pouring into a broadcaster's uploaded form follows the form, not the
    shoot: a Japanese template expects 名前スーパー whatever was spoken."""
    from app import lang
    from app.agents import telop_sheet

    assert telop_sheet._TYPE_JA is lang.SHEET_TELOP_TYPE[lang.JA]
    assert telop_sheet._TYPE_JA["name"] == "名前スーパー"
